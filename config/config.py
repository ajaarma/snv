#!/usr/bin/python

from xml.etree import ElementTree
from xml.dom import minidom
from xml.etree.ElementTree import Element, SubElement, Comment, tostring
import xmltodict
import argparse
import re,sys,os
from collections import OrderedDict

#import xlwt
#import xlrd
from os.path import isfile,join 
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell

######################################################################################
#
# Description: Generic class for creating XML based configuration files; 
#
######################################################################################

class CONFIG:

    def __init__(self, elements=[]):
        self.__elements={}
        for e in elements:
            self.__elemets[e]=1


    def display(self):
        print "Inside CONFIG class. Creating a configuration file for the NGC-SNV pipeline"

    def str2bool(self,v):
        ''' Utility function to convert string/unicode objects ('yes','true','t','1') 
        of boolean type (True/False) '''
        return v.lower() in ('yes','true','t','1')

    def processSNVArguments(self):
        ''' Subroutine to process command line argument for SNV pipeline '''

        cmdDict = {}

        parser = argparse.ArgumentParser(
                 description = 'python processSNV.py -a <xml-file> -p <proj-name>\n'
                               '-l <email-id> -g <gender-file> -e <exp-type> \n'
                               '-d <data-file-fof> -w <work-dir> -m <manifest-file> \n'
                               '-u <launch-flag>'
                 )
        parser.add_argument('-a','--analXML',help='Analysis XML file: Analysis.xml',
                            action='store',dest='xmlFile',required=True)
        parser.add_argument('-p','--projDate',help='Project date (YYYYMMDD): 20191231',
                            action='store',dest='projDate',required=True)
        parser.add_argument('-m','--manifest',help='Manifest file:',
                            action='store',dest='manifestFile',required=True)
        parser.add_argument('-e','--analType',help='Analysis type: gvcfGT/gel/exeter',
                            action='store',dest='expType',required=True)
        parser.add_argument('-w','--workDir',help='Working directory',
                            action='store',dest='workDir',required=True)
        parser.add_argument('-g','--gender',help='Gender file: gender_genomic.txt',
                            action='store',dest='genderFile',required=True)
        parser.add_argument('-d','--dataFile',help='All samples data file: data-file.fof',
                            action='store',dest='dataFile',required=True)
        parser.add_argument('-l','--email',help='User email id',
                            action='store',dest='emailID',required=False)
        parser.add_argument('-s','--flaggedSamples',help='List of flagged samples',
                            action='store',dest='fsFile',required=False) 
        parser.add_argument('-r','--header',help='header Info',
                            action='store',dest='header',required=False)
        parser.add_argument('-f','--famFile',help='List of family ids',
                            action='store',dest='famFile',required=True)
        parser.add_argument('-u','--launch',help='List of family ids',
                            action='store',dest='launchFlag',required=False)        
        parser.add_argument('-v','--version',help='Show programs version and exit\n\n\n',
                            action='version',version='%(prog)s 1.0')

        self.results = parser.parse_args()
        self.manifest_file = self.results.manifestFile
        self.xml_file = os.path.abspath(self.results.xmlFile)
        self.proj_date = self.results.projDate
        self.anal_type = self.results.expType
        self.email_id = self.results.emailID
        self.gender_file = self.results.genderFile
        self.work_dir = os.path.abspath(self.results.workDir)
        self.fam_file = os.path.abspath(self.results.famFile)
        self.data_file = os.path.abspath(self.results.dataFile)
        self.launch_flag = self.results.launchFlag
        
        try:
            self.flag_samples = self.results.fsFile
        except:
            self.flag_samples=''

        try:
            self.head_info = self.results.header
        except:
            self.head_info = ''

        
        cmd_dict = {'manifest':self.manifest_file,'xml':self.xml_file,
                    'proj':self.proj_date,'analType':self.anal_type,
                    'workDir':self.work_dir,'fam':self.fam_file,
                    'allSample':self.data_file,'flagSample':self.flag_samples,
                    'header':self.head_info,'gender':self.gender_file,
                    'email':self.email_id,'launch':self.launch_flag
                   }


        return cmd_dict


    def processXTRArguments(self):

        ''' Subroutine to process command line arguments for prerprocessing Exeter
        vcf files'''

        cmd_dict = {}
        parser = argparse.ArgumentParser(
                 description = 'python processXTR.py -a <xml-file> \n' 
                                '-i <exter-files> -d <inp-vcf-dir> \n'
                                '-o <out-dir> -e <anal-type>\n'
                 )
        parser.add_argument('-a','--xmlFile',help='Analysis-XML file',
                            action='store',dest='xmlFile',required=True) 
        parser.add_argument('-i','--inpFile',help='Trio ID mapping file',
                            action='store',dest='inpFile',required=True) 
        parser.add_argument('-d','--inpVCF',help='Exeter VCF directory',
                            action='store',dest='vcfDir',required=True)
        parser.add_argument('-o','--outDir',help='Output directory',
                            action='store',dest='outDir',required=True)
        parser.add_argument('-e','--expType',help='Analysis type',
                            action='store',dest='expType',required=True)

        self.results = parser.parse_args()
        self.xml_file = os.path.abspath(self.results.xmlFile)
        self.xtr_map_file = os.path.abspath(self.results.inpFile)
        self.vcf_dir = os.path.abspath(self.results.vcfDir)
        self.out_dir = os.path.abspath(self.results.outDir)
        self.exp_type = self.results.expType

        cmd_dict = {'xtrMap':self.xtr_map_file,'vcfDir':self.vcf_dir,
                    'outDir':self.out_dir,'expType':self.exp_type,
                    'xml':self.xml_file
                   }

        return cmd_dict

    def processCountVarArguments(self):
        ''' Subroutine to process Counting of variants '''

        cmd_dict = {}
        parser = argparse.ArgumentParser(
                             description = 'python countVariants.py -i <inp-dir> \n'
                                           '-o <out-dir> -a <anal-type>\n'
                             )

        parser.add_argument('-i','--inpDir',help='input VCF directory',
                            action='store',dest='inpDir',required=True)
        parser.add_argument('-o','--outDir',help='output directory',
                            action='store',dest='outDir',required=True)
        parser.add_argument('-a','--xmlFile',help='Analysis type',
                            action='store',dest='xmlFile',required=True)
        parser.add_argument('-p','--projDate',help='Project date (YYYYMMDD): 20191231',
                            action='store',dest='projDate',required=True)

        self.results = parser.parse_args()
        self.inp_dir = self.results.inpDir
        self.out_dir = self.results.outDir
        self.xml_file = self.results.xmlFile
        self.proj_date = self.results.projDate

        cmd_dict = {'inpDir':self.inp_dir,'out_dir':self.out_dir,'analType':self.xml_file,
                    'proj':self.proj_date}

        return cmd_dict

    def processXMLArguments(self):
        ''' Subroutine to process user configuration file '''
        
        cmd_dict = {}
        parser = argparse.ArgumentParser(
                     description = 'python createAnalysisXML.py -u <user-config> \n'
                                           '-b <base-xml-file> -o <out-xml-file> \n'
                    )

        parser.add_argument('-u','--userConfig',help='User specified config file',
                                    action='store',dest='userConfig',required=True)
        parser.add_argument('-b','--baseXml',help='Base XML config file',
                                    action='store',dest='baseXML',required=True)
        parser.add_argument('-o','--outXml',help='Output XML file',
                                    action='store',dest='outXML',required=True)

        self.results = parser.parse_args()
        self.user_config = os.path.abspath(self.results.userConfig)
        self.base_xml = os.path.abspath(self.results.baseXML)
        self.out_xml = os.path.abspath(self.results.outXML)

        cmd_dict = {'userConfig':self.user_config,'baseXml':self.base_xml,'outXml':self.out_xml}

        return cmd_dict

    
    def processInit(self,work_dir,cmd_args, proj_date):

        """ Creating temporary directories inside the work directory """
        
        print "Creating temporary directories inside the work directory \n"

        self.work_dir = work_dir
        self.date_str = proj_date
        self.tmp_dir = os.path.abspath(self.work_dir+'/'+self.date_str)
        self.tmp_bin = os.path.abspath(self.tmp_dir+'/tmp_binaries')
        self.tmp_data = os.path.abspath(self.tmp_dir+'/tmp_data')
        self.tmp_status = os.path.abspath(self.tmp_dir+'/tmp_status')
        self.tmp_log = os.path.abspath(self.tmp_dir+'/tmp_log')
        self.sb_log = os.path.abspath(self.tmp_dir+'/sb_log')
    
        self.tmp_dict = {'tmpDir':self.tmp_dir,'tmpBin':self.tmp_bin,'tmpData':self.tmp_data,
                    'tmpStat':self.tmp_status,'tmpLog':self.tmp_log,'sbLog':self.sb_log
                    }

        if os.path.isdir(self.work_dir+'/'+self.date_str):

            print ' -- The entered project space already exists. Do you want to delete it?'
            self.text = raw_input('Enter y/n: ')
            print self.text

            if self.text.lower() == 'y':
                print ' -- Deleting existing project. Creating fresh workspace with name: ',self.date_str
                os.system('rm -rf '+self.work_dir+'/'+self.date_str)

                for keys,values in self.tmp_dict.items():
                    os.system('mkdir -p '+values)

            elif self.text.lower() == 'n':
                pass
        else:
            for keys,values in self.tmp_dict.items():
                os.system('mkdir -p '+values)

        wh = open(self.tmp_dir+"/Command_Entered.txt","a+")
        print >>wh,"python "+" ".join(cmd_args)
        wh.close()

        return self.tmp_dict


    def createManifestFile(self,xtr_map_file,vcf_dir,out_dir,mh):
        
        ''' Subroutine for creating manifest file for any given mapping file '''

        self.vcfDict = OrderedDict()
        self.genderDict = OrderedDict()

        fh = open(xtr_map_file)

        for lines in fh:
            lines = lines.strip()
            strs = re.split('\t',lines);strs=[x.rstrip() for x in strs if x]
            #print strs 
            self.fam_id = strs[0]
            self.proband = strs[1]
            self.mother = strs[2]
            self.father = strs[3]
           
            self.proband_gender = ''

            if re.search('^male',strs[4],re.IGNORECASE):
                self.proband_gender = 'M'
            elif re.search('^female',strs[4],re.IGNORECASE):
                self.proband_gender = 'F'

            
            self.affected = '2'
            self.ilmn_id = 'None'
            self.fam_size = 'trio'
            self.load_hpc_date = '2019-12-19'
            self.bam_path = 'None'

            self.vcf_file =  vcf_dir+'/'+self.fam_id+'.vcf'
            self.out_vcf_gz = out_dir+'/'+self.fam_id+'.vcf.gz'
            
            #self.vcfDict[self.vcf_file] = 1 
            
            if self.proband !='Proband' and self.mother != 'Mother' \
                                        and self.father != 'Father':

                self.vcfDict[self.vcf_file] = 1 
                
                self.genderDict[self.proband] = self.proband_gender
                self.genderDict[self.mother] = 'F'
                self.genderDict[self.father] = 'M'

                proband_list = [self.fam_id,self.proband,self.father,self.mother,
                                self.proband_gender,self.affected,self.ilmn_id,
                                self.fam_size,self.load_hpc_date,self.bam_path,
                                self.out_vcf_gz]
                mother_list = [self.fam_id,self.mother,'None','None',
                                'F','1',self.ilmn_id,
                                self.fam_size,self.load_hpc_date,self.bam_path,
                                'None']
                father_list = [self.fam_id,self.father,'None','None',
                                'M','1',self.ilmn_id,
                                self.fam_size,self.load_hpc_date,self.bam_path,
                                'None']
   
                if not re.search('^File name',self.fam_id):
                    print >>mh,'\t'.join(proband_list)
                    print >>mh,'\t'.join(mother_list)
                    print >>mh,'\t'.join(father_list)
            
        return self.genderDict,self.vcfDict,mh

    def prettify(self,elem):

        rough_string = ElementTree.tostring(elem,'utf-8')
        reparsed = minidom.parseString(rough_string)
        return reparsed.toprettyxml(indent="  ")

    def getConfigTop(self, version_type):

        top = Element('config')
        comment = Comment("Configuration file for the NGC-SNV pipeline")
        top.append(comment)
        version = SubElement(top,"version")
        version.text = version_type

        return top

    def getGeneral(self):
        pass
  
    def getConfigDict(self, inp_file,module_name=[]):

        with open(inp_file) as fd:
            doc = xmltodict.parse(fd.read())

        doc = doc["config"]
        #print "The item classes present in Configuration files are: "+",".join(doc.keys())

        if module_name:
            config_dict = doc[module_name]
        else:
            config_dict = doc


        return config_dict

    def getUserConfigDict(self,userConfigFile):
        ''' Subroutine to read process the User defined configuration file and
            return as dictionary object 
        '''

        fh = open(userConfigFile)

        for lines in fh:
            lines = lines.strip()
            strs = re.split('\n',lines)
            
            if re.search('\=',strs[0]) and strs[0] !='':
                pass
            elif strs[0]:
                print 'The input user configuration file is not in correct format'+\
                        'Please read the online documentation '
                print 'The faulty line is: ',lines
                sys.exit()
        
        fh.close()
        
        userConfigDict = OrderedDict()

        fh = open(userConfigFile)
        for lines in fh:
            lines = lines.strip()
            strs = re.split('\=',lines);strs=[x.strip() for x in strs]
            if strs[0] and strs[1]:
                userConfigDict[strs[0]] = strs[1]
        
        fh.close()

        return userConfigDict

    def updateBaseXML(self,baseConfigDict,userConfigDict):
        ''' Subroutine to update the Base XML file with user specified 
            configuration file 
        '''

        for user_key in userConfigDict:
            if re.search('genomeBuild',user_key):
                try:
                    baseConfigDict['general'][user_key] = userConfigDict[user_key]
                except KeyError:
                    print 'The corresponding xml-field: ',user_key,' doesnot exist'
                    print 'Please enter correct xml-field tag: genomeBuild'
                
            if re.search('resourceDir',user_key):
                try:
                    baseConfigDict['general'][user_key] = userConfigDict[user_key]
                except KeyError:
                    print 'The corresponding xml-field: ',user_key,' doesnot exist'
                    print 'Please enter correct xml-field tag: resourceDir'
            if re.search('refFasta',user_key):
                try:
                    baseConfigDict['general'][user_key] = userConfigDict[user_key]
                except KeyError:
                    print 'The corresponding xml-field: ',user_key,' doesnot exist'
                    print 'Please enter correct xml-field tag: refFasta'
            
            if re.search('gnomad_g',user_key):
                try:
                    baseConfigDict['annotation']['cust-annot']['GNOMADg']['vcf'] = userConfigDict[user_key]
                except KeyError:
                    print 'The corresponding xml-field: ',user_key,' doesnot exist'
                    print 'Please enter correct xml-field tag: GNOMADg'
                 
            if re.search('gnomad_e',user_key):
                try:
                    baseConfigDict['annotation']['cust-annot']['GNOMADe']['vcf'] = userConfigDict[user_key]
                except KeyError:
                    print 'The corresponding xml-field: ',user_key,' doesnot exist'
                    print 'Please enter correct xml-field tag: GNOMADe'

            if re.search('dir_cache',user_key):
                try:
                    baseConfigDict['annotation']['defPar'][user_key] = userConfigDict[user_key]
                except KeyError:
                    print 'The corresponding xml-field: ',user_key,' doesnot exist'
                    print 'Please enter correct xml-field tag: <dir_cache>'

            if re.search('^exac$',user_key):
                try:
                    baseConfigDict['annotation']['cust-annot']['EXAC']['vcf'] = userConfigDict[user_key]
                except KeyError:
                    print 'The corresponding xml-field: ',user_key,' doesnot exist'
                    print 'Please enter correct xml-field tag: <EXAC>'
  
            if re.search('^exac_t$',user_key):
                try:
                    baseConfigDict['annotation']['cust-annot']['EXACt']['vcf'] = userConfigDict[user_key]
                except KeyError:
                    print 'The corresponding xml-field: ',user_key,' doesnot exist'
                    print 'Please enter correct xml-field tag: <exac_t>'
 
            if re.search('^clinvar$',user_key):
                try:
                    baseConfigDict['annotation']['cust-annot']['CLINVAR']['vcf'] = userConfigDict[user_key]
                except KeyError:
                    print 'The corresponding xml-field: ',user_key,' doesnot exist'
                    print 'Please enter correct xml-field tag: <CLINVAR>'
 
            if re.search('^hgmd$',user_key):
                try:
                    baseConfigDict['annotation']['cust-annot']['HGMD']['vcf'] = userConfigDict[user_key]
                except KeyError:
                    print 'The corresponding xml-field: ',user_key,' doesnot exist'
                    print 'Please enter correct xml-field tag: <HGMD>'
 
            if re.search('^cadd_snv$',user_key):
                try:
                    baseConfigDict['annotation']['plugIn']['CADD']['cadd_snv'] = userConfigDict[user_key]
                except KeyError:
                    print 'The corresponding xml-field: ',user_key,' doesnot exist'
                    print 'Please enter correct xml-field tag: <cadd_snv>'
 
            if re.search('^cadd_indel$',user_key):
                try:
                    baseConfigDict['annotation']['plugIn']['CADD']['cadd_indel'] = userConfigDict[user_key]
                except KeyError:
                    print 'The corresponding xml-field: ',user_key,' doesnot exist'
                    print 'Please enter correct xml-field tag: <cadd_indel>'
 
            if re.search('^exac_pli$',user_key):
                try:
                    baseConfigDict['annotation']['plugIn']['ExACpLI'] = userConfigDict[user_key]
                except KeyError:
                    print 'The corresponding xml-field: ',user_key,' doesnot exist'
                    print 'Please enter correct xml-field tag: <ExACpLI>'
 
            if re.search('^revel$',user_key):
                try:
                    baseConfigDict['annotation']['plugIn']['REVEL'] = userConfigDict[user_key]
                except KeyError:
                    print 'The corresponding xml-field: ',user_key,' doesnot exist'
                    print 'Please enter correct xml-field tag: <REVEL>'
 
            if re.search('^region_exons$',user_key):
                try:
                    baseConfigDict['regions']['exons'] = userConfigDict[user_key]
                except KeyError:
                    print 'The corresponding xml-field: ',user_key,' doesnot exist'
                    print 'Please enter correct xml-field tag: <exons>'
 
            if re.search('^region_par$',user_key):
                try:
                    baseConfigDict['regions']['PAR'] = userConfigDict[user_key]
                except KeyError:
                    print 'The corresponding xml-field: ',user_key,' doesnot exist'
                    print 'Please enter correct xml-field tag: <PAR>'

            if re.search('^ensembl$',user_key):
                try:
                    baseConfigDict['varPrior']['params']['l'] = userConfigDict[user_key]
                except KeyError:
                    print 'The corresponding xml-field: ',user_key,' doesnot exist'
                    print 'Please enter correct xml-field tag: <ensembl>'

            if re.search('^hpo$',user_key):
                try:
                    baseConfigDict['famFilter'][user_key] = userConfigDict[user_key]
                except KeyError:
                    print 'The corresponding xml-field: ',user_key,' doesnot exist'
                    print 'Please enter correct xml-field tag: <hpo>'

            if re.search('^phenotypes$',user_key):
                try:
                    baseConfigDict['famFilter'][user_key] = userConfigDict[user_key]
                except KeyError:
                    print 'The corresponding xml-field: ',user_key,' doesnot exist'
                    print 'Please enter correct xml-field tag: <phenotypes>'


        return baseConfigDict

    def convertTxt2XL(self,inpDir):
        ''' Utility function to convert list of Text files to Excel sheets '''

        #textfiles = [join(inpDir,f) for f in os.listdir(inpdir) 
        #                        if isfile(join(inpDir,f,'filtering')) and '.txt' in f ]

        wb = Workbook()

        for fam_id in os.listdir(inpDir):
            #fam_id = 'TwEx_EX1912835-TwEx_EX1912836-TwEx_EX1912837'
            try:
                txt_file = os.listdir(join(inpDir,fam_id,'filtering'))
                #print fam_id,'\t',fam_id[0]
                #print txt_file
                proband_id = re.split('\-',fam_id)[0]
                path_txt_file = join(inpDir,fam_id,'filtering',txt_file[0])
                #worksheet = workbook.add_sheet(str(proband_id))

                f = open(path_txt_file,'r+')
                data = f.readlines()
               
                proband_id = fam_id
                wb.create_sheet(proband_id)
                this_sheet = wb[proband_id]
                for row in data:
                    this_sheet.append(row.split('\t'))
                f.close()

            except IndexError:
                print fam_id

        wb.save(join(inpDir,'exeter_exomes_vars_famID.xlsx'))
